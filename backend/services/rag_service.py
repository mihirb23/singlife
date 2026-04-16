# rag_service.py — vector store with incremental indexing
# scans knowledge_base/ for all supported file types (.txt, .pdf, .xlsx, .json),
# extracts text, chunks it, and embeds into chromadb.
# uses a manifest to track indexed state — only re-embeds new/changed files.

import hashlib
import json
import logging
import re
from datetime import datetime
from difflib import SequenceMatcher
from pathlib import Path
from typing import Dict, List, Optional

import chromadb
from chromadb.utils import embedding_functions

logger = logging.getLogger(__name__)

KB_DIR = Path(__file__).parent.parent.parent / 'knowledge_base'
CHROMA_DIR = Path(__file__).parent.parent.parent / 'chroma_db'
MANIFEST_PATH = CHROMA_DIR / 'index_manifest.json'

CHUNK_SIZE = 1500
CHUNK_OVERLAP = 200

SUPPORTED_EXTENSIONS = {'.txt', '.pdf', '.xlsx', '.json'}

PARSER_VERSIONS = {
    '.txt': 1,
    '.pdf': 1,
    '.xlsx': 1,
    '.json': 1,
}


# ---------------------------------------------------------------------------
# Chunking
# ---------------------------------------------------------------------------

def chunk_text(text: str, chunk_size: int = CHUNK_SIZE, overlap: int = CHUNK_OVERLAP) -> List[str]:
    """Split text into overlapping chunks so boundary-spanning rules
    still get captured in at least one chunk."""
    chunks = []
    start = 0
    while start < len(text):
        end = start + chunk_size
        chunk = text[start:end].strip()
        if chunk:
            chunks.append(chunk)
        start += chunk_size - overlap
    return chunks


# ---------------------------------------------------------------------------
# File hashing
# ---------------------------------------------------------------------------

def file_content_hash(path: Path) -> str:
    """SHA-256 of raw file bytes — used for change detection."""
    h = hashlib.sha256()
    with open(path, 'rb') as f:
        for block in iter(lambda: f.read(8192), b''):
            h.update(block)
    return h.hexdigest()


# ---------------------------------------------------------------------------
# Text extractors — one per supported file type
# ---------------------------------------------------------------------------

def extract_txt(path: Path) -> str:
    return path.read_text(encoding='utf-8', errors='replace').strip()


def extract_json(path: Path) -> str:
    text = path.read_text(encoding='utf-8').strip()
    try:
        data = json.loads(text)
        return json.dumps(data, indent=2, ensure_ascii=False)
    except json.JSONDecodeError:
        return text


def extract_pdf(path: Path) -> str:
    import pypdf
    reader = pypdf.PdfReader(str(path))
    pages = []
    for i, page in enumerate(reader.pages):
        text = page.extract_text() or ""
        if text.strip():
            pages.append(f"--- Page {i + 1} ---\n{text.strip()}")
    return "\n\n".join(pages)


def extract_xlsx(path: Path) -> str:
    import openpyxl
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    sections = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows: List[str] = []
        headers = None
        for row in ws.iter_rows(values_only=True):
            values = [str(cell) if cell is not None else "" for cell in row]
            if all(v == "" for v in values):
                continue
            if headers is None:
                headers = values
                rows.append(" | ".join(headers))
                rows.append("-" * len(rows[-1]))
            else:
                rows.append(" | ".join(values))
        if rows:
            sections.append(f"Sheet: {sheet_name}\n" + "\n".join(rows))
    wb.close()
    return "\n\n".join(sections)


EXTRACTORS = {
    '.txt': extract_txt,
    '.pdf': extract_pdf,
    '.xlsx': extract_xlsx,
    '.json': extract_json,
}


def extract_text(path: Path) -> Optional[str]:
    """High-level convenience: dispatch to the right extractor by extension."""
    extractor = EXTRACTORS.get(path.suffix.lower())
    if not extractor:
        return None
    try:
        return extractor(path)
    except Exception as e:
        logger.error(f"Extraction failed for {path.name}: {e}")
        return None


# ---------------------------------------------------------------------------
# Manifest — tracks which files are indexed and their content hashes
# ---------------------------------------------------------------------------

def load_manifest() -> dict:
    if MANIFEST_PATH.exists():
        try:
            return json.loads(MANIFEST_PATH.read_text(encoding='utf-8'))
        except Exception as e:
            logger.warning(f"Corrupt manifest, starting fresh: {e}")
    return {"manifest_version": 1, "files": {}}


def save_manifest(manifest: dict):
    CHROMA_DIR.mkdir(parents=True, exist_ok=True)
    MANIFEST_PATH.write_text(json.dumps(manifest, indent=2), encoding='utf-8')


def scan_kb_files() -> Dict[str, Path]:
    """Return {filename: Path} for every supported file in knowledge_base/."""
    KB_DIR.mkdir(exist_ok=True)
    found = {}
    for f in sorted(KB_DIR.iterdir()):
        if f.is_file() and f.suffix.lower() in SUPPORTED_EXTENSIONS:
            found[f.name] = f
    return found


def compute_diff(current_files: Dict[str, Path], manifest: dict) -> dict:
    """Compare current KB files against manifest.
    Returns lists of new / changed / unchanged / deleted entries."""
    indexed = manifest.get("files", {})

    new_files = []
    changed_files = []
    unchanged_files = []
    deleted_files = []

    for name, path in current_files.items():
        content_hash = file_content_hash(path)
        ext = path.suffix.lower()
        parser_ver = PARSER_VERSIONS.get(ext, 1)

        if name not in indexed:
            new_files.append({"name": name, "path": path, "hash": content_hash})
        else:
            entry = indexed[name]
            hash_changed = entry.get("content_hash") != content_hash
            parser_changed = entry.get("parser_version") != parser_ver
            if hash_changed or parser_changed:
                changed_files.append({
                    "name": name,
                    "path": path,
                    "hash": content_hash,
                    "old_chunk_ids": entry.get("chunk_ids", []),
                })
            else:
                unchanged_files.append(name)

    for name in indexed:
        if name not in current_files:
            deleted_files.append({
                "name": name,
                "old_chunk_ids": indexed[name].get("chunk_ids", []),
            })

    return {
        "new": new_files,
        "changed": changed_files,
        "unchanged": unchanged_files,
        "deleted": deleted_files,
    }


# ---------------------------------------------------------------------------
# VectorStore — manages chromadb collection with incremental indexing
# ---------------------------------------------------------------------------

class VectorStore:
    """Chromadb-backed vector store with manifest-driven incremental indexing.
    Only re-embeds new or changed KB files; removes stale chunks for deleted
    or updated files."""

    def __init__(self):
        self.client = None
        self.collection = None
        self._init_store()

    def _init_store(self):
        try:
            CHROMA_DIR.mkdir(parents=True, exist_ok=True)
            self.client = chromadb.PersistentClient(path=str(CHROMA_DIR))

            ef = embedding_functions.SentenceTransformerEmbeddingFunction(
                model_name="all-MiniLM-L6-v2"
            )

            self.collection = self.client.get_or_create_collection(
                name="singlife_kb",
                embedding_function=ef,
                metadata={"hnsw:space": "cosine"},
            )

            logger.info(f"Vector store ready — {self.collection.count()} chunks indexed")
        except Exception as e:
            logger.error(f"Failed to init vector store: {e}")
            self.collection = None

    # ------------------------------------------------------------------
    # Incremental indexing
    # ------------------------------------------------------------------

    def index_documents(self):
        """Scan KB, diff against manifest, embed only what changed."""
        if not self.collection:
            logger.error("Collection not available — can't index")
            return

        fresh_manifest = not MANIFEST_PATH.exists()
        manifest = load_manifest()

        # First run after migrating from old full-wipe system: clear orphans
        if fresh_manifest and self.collection.count() > 0:
            logger.info("No manifest found — clearing legacy index for clean re-build")
            existing_ids = self.collection.get()["ids"]
            if existing_ids:
                self._batch_delete(existing_ids)
            manifest = {"manifest_version": 1, "files": {}}

        current_files = scan_kb_files()
        diff = compute_diff(current_files, manifest)

        n_new = len(diff["new"])
        n_chg = len(diff["changed"])
        n_del = len(diff["deleted"])
        n_unch = len(diff["unchanged"])

        if n_new == 0 and n_chg == 0 and n_del == 0:
            logger.info(f"Incremental index: nothing to do ({n_unch} files unchanged)")
            return

        logger.info(f"Incremental index: {n_new} new, {n_chg} changed, "
                     f"{n_del} deleted, {n_unch} unchanged")

        # --- Remove stale chunks for changed / deleted files ---
        stale_ids = []
        for entry in diff["deleted"] + diff["changed"]:
            stale_ids.extend(entry.get("old_chunk_ids", []))
        if stale_ids:
            self._batch_delete(stale_ids)
            logger.info(f"Removed {len(stale_ids)} stale chunks")

        for entry in diff["deleted"]:
            manifest["files"].pop(entry["name"], None)
            logger.info(f"De-indexed: {entry['name']}")

        # --- Index new + changed files ---
        for entry in diff["new"] + diff["changed"]:
            name = entry["name"]
            path = entry["path"]
            content_hash = entry["hash"]
            ext = path.suffix.lower()
            hash_short = content_hash[:12]
            stat = path.stat()

            extractor = EXTRACTORS.get(ext)
            if not extractor:
                manifest["files"][name] = self._manifest_entry(
                    content_hash, stat, ext, chunk_ids=[],
                    status="unsupported", error=f"No extractor for {ext}")
                continue

            try:
                text = extractor(path)
                if not text or not text.strip():
                    manifest["files"][name] = self._manifest_entry(
                        content_hash, stat, ext, chunk_ids=[],
                        status="empty", error="No text extracted")
                    logger.warning(f"No text extracted from {name}")
                    continue

                chunks = chunk_text(text)
                chunk_ids = [f"{path.stem}_{hash_short}_{i}" for i in range(len(chunks))]
                metadatas = [
                    {"source": name, "chunk_index": i, "total_chunks": len(chunks)}
                    for i in range(len(chunks))
                ]

                self._batch_add(chunks, chunk_ids, metadatas)

                manifest["files"][name] = self._manifest_entry(
                    content_hash, stat, ext, chunk_ids=chunk_ids,
                    status="indexed")
                logger.info(f"Indexed {name}: {len(chunks)} chunks")

            except Exception as e:
                manifest["files"][name] = self._manifest_entry(
                    content_hash, stat, ext, chunk_ids=[],
                    status="error", error=str(e))
                logger.error(f"Failed to index {name}: {e}")

        save_manifest(manifest)
        logger.info(f"Manifest saved — {len(manifest['files'])} files tracked, "
                     f"{self.collection.count()} total chunks in store")

    # ------------------------------------------------------------------
    # Query
    # ------------------------------------------------------------------

    def query(self, question: str, top_k: int = 8, priority_sources: Optional[List[str]] = None) -> List[dict]:
        if not self.collection or self.collection.count() == 0:
            return []
        try:
            fetch_k = min(top_k * 3, self.collection.count()) if priority_sources else min(top_k, self.collection.count())
            results = self.collection.query(
                query_texts=[question],
                n_results=fetch_k,
            )
            chunks = []
            metadata_rows = results.get("metadatas") or []
            metadata_row = metadata_rows[0] if metadata_rows else []
            if metadata_row is None:
                metadata_row = []
            for i, doc in enumerate(results["documents"][0]):
                metadata = metadata_row[i] if i < len(metadata_row) and metadata_row[i] else {}
                chunks.append({
                    "text": doc,
                    "source": metadata.get("source"),
                    "chunk_index": metadata.get("chunk_index"),
                    "total_chunks": metadata.get("total_chunks"),
                    "score": results["distances"][0][i] if results["distances"] else None,
                })
            if priority_sources:
                wanted = {s.lower() for s in priority_sources}
                chunks.sort(
                    key=lambda c: (
                        0 if str(c.get("source", "")).lower() in wanted else 1,
                        c["score"] if c.get("score") is not None else 999999.0,
                    )
                )
                return chunks[:top_k]
            return chunks
        except Exception as e:
            logger.error(f"Vector store query failed: {e}")
            return []

    def resolve_source_by_name(self, query: str, confidence_min: float = 0.72, tie_margin: float = 0.03) -> dict:
        """Resolve a KB source from user query with deterministic ranking."""
        candidates = list(scan_kb_files().keys())
        if not candidates:
            return {
                "matched": False,
                "source": None,
                "candidates": [],
                "confidence": 0.0,
                "reason": "no_sources_available",
                "ambiguous": False,
            }

        query_lower = query.lower().strip()
        query_norm = self._normalize_token(query)

        scored = []
        for source in candidates:
            source_lower = source.lower()
            stem_lower = Path(source).stem.lower()
            source_norm = self._normalize_token(source_lower)
            stem_norm = self._normalize_token(stem_lower)

            score = 0.0
            reason = "fuzzy"
            if source_lower in query_lower:
                score = 1.0
                reason = "exact_filename"
            elif stem_lower and stem_lower in query_lower:
                score = 0.96
                reason = "exact_stem"
            elif source_norm and source_norm in query_norm:
                score = 0.93
                reason = "normalized_exact"
            elif stem_norm and stem_norm in query_norm:
                score = 0.90
                reason = "normalized_stem"
            else:
                score = max(
                    SequenceMatcher(None, query_norm, source_norm).ratio(),
                    SequenceMatcher(None, query_norm, stem_norm).ratio(),
                )

            scored.append({
                "source": source,
                "score": float(score),
                "reason": reason,
            })

        scored.sort(key=lambda item: item["score"], reverse=True)
        top = scored[0]
        runner_up = scored[1] if len(scored) > 1 else None

        if top["score"] < confidence_min:
            return {
                "matched": False,
                "source": None,
                "candidates": [s["source"] for s in scored[:3]],
                "confidence": top["score"],
                "reason": "below_confidence_threshold",
                "ambiguous": False,
            }

        if runner_up and runner_up["score"] >= confidence_min and abs(top["score"] - runner_up["score"]) <= tie_margin:
            return {
                "matched": False,
                "source": None,
                "candidates": [top["source"], runner_up["source"]],
                "confidence": top["score"],
                "reason": "ambiguous_match",
                "ambiguous": True,
            }

        return {
            "matched": True,
            "source": top["source"],
            "candidates": [s["source"] for s in scored[:3]],
            "confidence": top["score"],
            "reason": top["reason"],
            "ambiguous": False,
        }

    def get_full_source_text(self, source: str) -> dict:
        """Load extracted full text for a single source file."""
        path = KB_DIR / source
        if not path.exists() or not path.is_file():
            return {
                "source": source,
                "text": "",
                "char_count": 0,
                "ordered": True,
                "extractor_type": None,
                "error": "source_not_found",
            }

        text = extract_text(path)
        if not text:
            return {
                "source": source,
                "text": "",
                "char_count": 0,
                "ordered": True,
                "extractor_type": path.suffix.lower().lstrip("."),
                "error": "extraction_failed_or_empty",
            }

        return {
            "source": source,
            "text": text,
            "char_count": len(text),
            "ordered": True,
            "extractor_type": path.suffix.lower().lstrip("."),
            "error": None,
        }

    def get_chunks_for_source(self, source: str) -> List[dict]:
        """Return all chunks for a source in original chunk order."""
        if not self.collection:
            return []
        try:
            result = self.collection.get(
                where={"source": source},
                include=["documents", "metadatas"],
            )
            documents = result.get("documents", []) or []
            metadatas = result.get("metadatas", []) or []
            items = []
            for i, text in enumerate(documents):
                md = metadatas[i] if i < len(metadatas) and metadatas[i] else {}
                items.append({
                    "text": text,
                    "source": source,
                    "chunk_index": int(md.get("chunk_index", i)),
                    "total_chunks": int(md.get("total_chunks", len(documents))),
                })
            items.sort(key=lambda item: item["chunk_index"])
            return items
        except Exception as e:
            logger.error(f"Failed to load chunks for source {source}: {e}")
            return []

    def get_expanded_context(
        self,
        question: str,
        top_k: int = 8,
        priority_sources: Optional[List[str]] = None,
        neighbor_window: int = 1,
    ) -> dict:
        """Retrieve semantic hits and expand with neighboring chunks by source."""
        matched = self.query(question, top_k=top_k, priority_sources=priority_sources)
        if not matched:
            return {
                "chunks": [],
                "grouped_by_source": {},
                "coverage": {},
                "quality": {
                    "result_count": 0,
                    "best_score": None,
                    "unique_sources": 0,
                    "sufficient": False,
                    "reason": "no_chunks",
                },
            }

        grouped_hits: Dict[str, List[dict]] = {}
        for hit in matched:
            source = hit.get("source")
            if not source:
                continue
            grouped_hits.setdefault(source, []).append(hit)

        expanded_map: Dict[tuple, dict] = {}
        coverage = {}
        for source, hits in grouped_hits.items():
            source_chunks = self.get_chunks_for_source(source)
            if not source_chunks:
                continue
            chunk_by_idx = {c["chunk_index"]: c for c in source_chunks}
            matched_indices = sorted({
                int(h["chunk_index"]) for h in hits if h.get("chunk_index") is not None
            })
            selected_indices = set()
            for idx in matched_indices:
                for i in range(max(0, idx - neighbor_window), idx + neighbor_window + 1):
                    if i in chunk_by_idx:
                        selected_indices.add(i)

            for idx in sorted(selected_indices):
                item = dict(chunk_by_idx[idx])
                if idx in matched_indices:
                    hit = next((h for h in hits if h.get("chunk_index") == idx), None)
                    item["score"] = hit.get("score") if hit else None
                    item["is_match"] = True
                else:
                    item["score"] = None
                    item["is_match"] = False
                expanded_map[(source, idx)] = item

            total_chunks = source_chunks[0]["total_chunks"] if source_chunks else len(selected_indices)
            coverage[source] = {
                "matched_chunks": len(matched_indices),
                "expanded_chunks": len(selected_indices),
                "total_chunks": total_chunks,
            }

        expanded_chunks = list(expanded_map.values())
        expanded_chunks.sort(key=lambda c: (str(c.get("source", "")), int(c.get("chunk_index", 0))))

        scores = [c["score"] for c in matched if c.get("score") is not None]
        best_score = min(scores) if scores else None
        result_count = len(matched)
        unique_sources = len({c.get("source") for c in matched if c.get("source")})
        sufficient = result_count >= max(2, min(top_k, 4))
        reason = "ok"
        if best_score is not None and best_score > 1.15:
            sufficient = False
            reason = "weak_similarity_scores"
        elif unique_sources > max(3, top_k // 2):
            sufficient = False
            reason = "over_fragmented_sources"
        elif not sufficient:
            reason = "too_few_matches"

        grouped_by_source = {}
        for chunk in expanded_chunks:
            grouped_by_source.setdefault(chunk["source"], []).append(chunk)

        return {
            "chunks": expanded_chunks,
            "grouped_by_source": grouped_by_source,
            "coverage": coverage,
            "quality": {
                "result_count": result_count,
                "best_score": best_score,
                "unique_sources": unique_sources,
                "sufficient": sufficient,
                "reason": reason,
            },
        }

    def is_available(self) -> bool:
        return self.collection is not None and self.collection.count() > 0

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _manifest_entry(content_hash, stat, ext, *, chunk_ids,
                        status, error=None) -> dict:
        return {
            "content_hash": content_hash,
            "size": stat.st_size,
            "mtime": stat.st_mtime,
            "parser_type": ext,
            "parser_version": PARSER_VERSIONS.get(ext, 1),
            "chunk_ids": chunk_ids,
            "status": status,
            "error": error,
            "indexed_at": datetime.now().isoformat(),
        }

    def _batch_delete(self, ids: List[str]):
        batch = 100
        for i in range(0, len(ids), batch):
            try:
                self.collection.delete(ids=ids[i:i + batch])
            except Exception as e:
                logger.error(f"Failed to delete chunk batch: {e}")

    def _batch_add(self, documents: List[str], ids: List[str],
                   metadatas: List[dict]):
        batch = 100
        for i in range(0, len(documents), batch):
            self.collection.add(
                documents=documents[i:i + batch],
                ids=ids[i:i + batch],
                metadatas=metadatas[i:i + batch],
            )

    @staticmethod
    def _normalize_token(value: str) -> str:
        return re.sub(r"[^a-z0-9]+", "", (value or "").lower())
